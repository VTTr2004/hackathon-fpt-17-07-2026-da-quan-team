from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from app.schemas.common import Evidence

from .schemas import CashActivity, CashDirection, CashFlowDataset, CashFlowTransaction


HEADER_SCAN_ROWS = 50


def _text(value: Any) -> str:
    return str(value or "").strip().casefold()


def _header_index(headers: list[str], *names: str) -> int | None:
    return next(
        (index for index, header in enumerate(headers) if any(name in header for name in names)),
        None,
    )


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def _amount(value: Any) -> Decimal:
    try:
        amount = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError("amount must be numeric") from exc
    if not amount.is_finite() or amount < 0:
        raise ValueError("amount must be a finite non-negative number")
    return amount


def _period(value: Any) -> str | None:
    if isinstance(value, datetime | date):
        return value.strftime("%Y-%m")
    return None


def _find_cashbook_headers(worksheet: Any) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
        if row_number > HEADER_SCAN_ROWS:
            break
        headers = [_text(value) for value in row]
        date_index = _header_index(headers, "ngày", "ngay", "date")
        inflow_index = _header_index(headers, "tiền vào", "tien vao", "inflow")
        outflow_index = _header_index(headers, "tiền ra", "tien ra", "outflow")
        if None in {date_index, inflow_index, outflow_index}:
            continue
        description_index = _header_index(headers, "diễn giải", "dien giai", "description")
        category_index = _header_index(headers, "nhóm", "nhom", "category")
        candidates.append(
            {
                "header_row": row_number,
                "date_index": date_index,
                "inflow_index": inflow_index,
                "outflow_index": outflow_index,
                "description_index": description_index,
                "category_index": category_index,
                "is_detailed": description_index is not None or category_index is not None,
            }
        )
    return candidates


def _read_balances(worksheet: Any, warnings: list[str]) -> tuple[Decimal | None, Decimal | None]:
    opening_cash: Decimal | None = None
    reported_ending_cash: Decimal | None = None
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
        if row_number > HEADER_SCAN_ROWS:
            break
        label = _text(_cell(row, 0))
        if not label:
            continue
        value = next((cell for cell in row[1:] if cell is not None and str(cell).strip()), None)
        if value is None:
            continue
        try:
            if "số dư đầu kỳ" in label or "so du dau ky" in label:
                opening_cash = _amount(value)
            elif "số dư cuối kỳ" in label or "so du cuoi ky" in label:
                reported_ending_cash = _amount(value)
        except ValueError:
            warnings.append(f"Invalid balance in {worksheet.title} row {row_number}; value was ignored.")
    return opening_cash, reported_ending_cash


def extract_cash_flow_documents(
    documents: list[dict[str, Any]],
) -> tuple[CashFlowDataset | None, list[Evidence], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, [], ["Workbook extraction is unavailable because openpyxl is not installed."]

    transactions: list[CashFlowTransaction] = []
    evidence: list[Evidence] = []
    warnings: list[str] = []
    opening_cash: Decimal | None = None
    reported_ending_cash: Decimal | None = None

    for document in documents:
        path = Path(document.get("storage_path") or "")
        if path.suffix.lower() != ".xlsx" or not path.exists():
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            warnings.append(f"Could not read workbook {document.get('filename', path.name)}: {exc}")
            continue

        candidates: list[tuple[Any, dict[str, Any]]] = []
        for worksheet in workbook.worksheets:
            sheet_opening, sheet_ending = _read_balances(worksheet, warnings)
            opening_cash = opening_cash if opening_cash is not None else sheet_opening
            reported_ending_cash = reported_ending_cash if reported_ending_cash is not None else sheet_ending
            candidates.extend((worksheet, candidate) for candidate in _find_cashbook_headers(worksheet))

        detailed_candidates = [candidate for candidate in candidates if candidate[1]["is_detailed"]]
        selected_candidates = detailed_candidates or candidates
        if candidates and not detailed_candidates:
            warnings.append(
                f"Workbook {document.get('filename', path.name)} has no detailed cashbook sheet; using summary rows."
            )

        for worksheet, header in selected_candidates:
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header["header_row"] + 1, values_only=True),
                header["header_row"] + 1,
            ):
                raw_date = _cell(row, header["date_index"])
                period = _period(raw_date)
                if period is None:
                    if any(value is not None for value in row):
                        warnings.append(f"Skipped row without a valid date: {worksheet.title} row {row_number}.")
                    continue
                for direction, amount_index in (
                    (CashDirection.INFLOW, header["inflow_index"]),
                    (CashDirection.OUTFLOW, header["outflow_index"]),
                ):
                    raw_amount = _cell(row, amount_index)
                    if raw_amount in (None, 0, ""):
                        continue
                    try:
                        amount = _amount(raw_amount)
                    except ValueError:
                        warnings.append(f"Skipped invalid amount: {worksheet.title} row {row_number}.")
                        continue
                    evidence_id = f"{document['id']}:{worksheet.title}:{row_number}:{direction.value}"
                    transactions.append(
                        CashFlowTransaction(
                            period=period,
                            date=raw_date,
                            direction=direction,
                            activity=CashActivity.UNCLASSIFIED,
                            category=str(_cell(row, header["category_index"]) or "cashbook"),
                            description=str(_cell(row, header["description_index"]) or ""),
                            amount=amount,
                            document_id=document["id"],
                            filename=document["filename"],
                            sheet=worksheet.title,
                            row_number=row_number,
                            evidence_id=evidence_id,
                        )
                    )
                    evidence.append(
                        Evidence(
                            evidence_id=evidence_id,
                            source_type="uploaded_workbook",
                            title=f"{document['filename']} / {worksheet.title} row {row_number}",
                            document_id=document["id"],
                        )
                    )

    if not transactions:
        return None, evidence, warnings
    return (
        CashFlowDataset(
            opening_cash=opening_cash,
            reported_ending_cash=reported_ending_cash,
            transactions=transactions,
            source_type="cashbook",
        ),
        evidence,
        warnings,
    )
