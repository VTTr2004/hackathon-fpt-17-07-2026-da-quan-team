from datetime import date, datetime
from pathlib import Path
from typing import Any

from .ingestion_schemas import SheetRowSample, WorkbookSheetProfile

MAX_PROFILE_ROWS = 50
MAX_PROFILE_COLUMNS = 30
MAX_CELL_TEXT = 160


def _profile_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()[:MAX_CELL_TEXT]
    if isinstance(value, (int, float, bool)) or value is None:
        return value
    return str(value)[:MAX_CELL_TEXT]


def profile_cash_flow_workbooks(documents: list[dict[str, Any]]) -> tuple[list[WorkbookSheetProfile], list[str]]:
    """Create a bounded, calculation-free workbook view for the mapping agent."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["Workbook profiling is unavailable because openpyxl is not installed."]

    profiles: list[WorkbookSheetProfile] = []
    warnings: list[str] = []
    for document in documents:
        path = Path(document.get("storage_path") or "")
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            warnings.append(f"Could not profile workbook {document.get('filename', path.name)}: {exc}")
            continue
        try:
            for worksheet in workbook.worksheets:
                sampled_rows: list[SheetRowSample] = []
                if worksheet.max_row is None or worksheet.max_column is None:
                    worksheet.calculate_dimension(force=True)
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                if max_row == 0 or max_column == 0:
                    profiles.append(
                        WorkbookSheetProfile(
                            document_id=str(document.get("id", "")),
                            filename=str(document.get("filename") or path.name),
                            sheet=worksheet.title,
                            max_row=max_row,
                            max_column=max_column,
                        )
                    )
                    continue
                for row_number, row in enumerate(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=min(max_row, MAX_PROFILE_ROWS),
                        max_col=min(max_column, MAX_PROFILE_COLUMNS),
                        values_only=True,
                    ),
                    1,
                ):
                    values = [_profile_value(value) for value in row]
                    if any(value not in (None, "") for value in values):
                        sampled_rows.append(SheetRowSample(row_number=row_number, values=values))
                profiles.append(
                    WorkbookSheetProfile(
                        document_id=str(document.get("id", "")),
                        filename=str(document.get("filename") or path.name),
                        sheet=worksheet.title,
                        max_row=max_row,
                        max_column=max_column,
                        sampled_rows=sampled_rows,
                    )
                )
        finally:
            workbook.close()
    return profiles, warnings
