import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from unicodedata import normalize as unicode_normalize

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.modules.cash_flow.ingestion_schemas import (
    FieldProposal,
    IngestionToolName,
    IngestionToolRequest,
    ProposalSource,
    ToolExecutionResult,
)
from app.modules.cash_flow.schemas import CashActivity, CashDirection, CashFlowDataset, CashFlowTransaction
from app.schemas.common import Evidence

TOOL_VERSION = "1.0.0"
MAX_DATA_ROWS = 100_000

TOOL_CONTRACTS: dict[IngestionToolName, dict[str, Any]] = {
    IngestionToolName.NORMALIZE_CASHBOOK: {
        "description": "Chuẩn hóa bảng thu chi thành giao dịch cash flow canonical.",
        "required_columns": ["date", "inflow", "outflow"],
        "optional_columns": [
            "transaction_id",
            "type",
            "category",
            "description",
            "account",
            "source_ref",
        ],
        "methodology_reference": "docs/methodology.md#cash-flow-normalization",
        "deterministic": True,
    },
    IngestionToolName.EXTRACT_FINANCIAL_FACTS: {
        "description": "Trích xuất các cặp nhãn/giá trị tài chính đã được mapping.",
        "required_columns": ["label", "value"],
        "optional_columns": ["unit", "notes"],
        "methodology_reference": "docs/methodology.md#document-derived-facts",
        "deterministic": True,
    },
    IngestionToolName.SUMMARIZE_SALES: {
        "description": "Tổng hợp doanh thu theo kỳ/kênh/phương thức thanh toán bằng code.",
        "required_columns": ["date", "net_amount"],
        "optional_columns": ["quantity", "channel", "payment_method", "category", "order_id"],
        "methodology_reference": "docs/methodology.md#sales-to-cash-support",
        "deterministic": True,
    },
    IngestionToolName.SUMMARIZE_PURCHASES: {
        "description": "Tổng hợp mua hàng/chi phí và khoản chưa thanh toán bằng code.",
        "required_columns": ["date", "total_amount"],
        "optional_columns": ["category", "supplier", "payment_status", "vat_amount", "invoice_id"],
        "methodology_reference": "docs/methodology.md#cost-and-payables-support",
        "deterministic": True,
    },
}

ALLOWED_FACT_FIELDS = frozenset(
    {
        "opening_cash",
        "reported_ending_cash",
        "current_cash",
        "cash_as_of",
        "currency",
        "minimum_cash_buffer",
        "monthly_rent",
        "lease_deposit",
        "employee_count",
    }
)

FACT_ALIASES = {
    "so du dau ky": "opening_cash",
    "opening cash": "opening_cash",
    "beginning cash": "opening_cash",
    "so du cuoi ky": "reported_ending_cash",
    "ending cash": "reported_ending_cash",
    "current cash": "current_cash",
    "tien mat hien co": "current_cash",
    "minimum cash buffer": "minimum_cash_buffer",
    "gia thue/thang": "monthly_rent",
    "tien thue/thang": "monthly_rent",
    "monthly rent": "monthly_rent",
    "tien dat coc": "lease_deposit",
    "deposit": "lease_deposit",
    "nhan su": "employee_count",
    "employees": "employee_count",
    "tien te": "currency",
    "currency": "currency",
}


def ingestion_tool_catalog() -> list[dict[str, Any]]:
    return [
        {"name": name.value, "version": TOOL_VERSION, "input_schema": contract}
        for name, contract in TOOL_CONTRACTS.items()
    ]


def _plain_text(value: Any) -> str:
    return str(value or "").strip()


def _normalized_text(value: Any) -> str:
    raw = _plain_text(value).replace("đ", "d").replace("Đ", "D")
    text = unicode_normalize("NFKD", raw).encode("ascii", "ignore").decode().casefold()
    return " ".join(text.split())


def _decimal(value: Any, *, field: str) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"{field} must be numeric") from exc
    if not number.is_finite() or number < 0:
        raise ValueError(f"{field} must be finite and non-negative")
    return number


def _period(value: Any) -> str:
    if isinstance(value, datetime | date):
        return value.strftime("%Y-%m")
    text = _plain_text(value)
    if len(text) >= 7 and text[:4].isdigit() and text[4] in {"-", "/"} and text[5:7].isdigit():
        month = int(text[5:7])
        if 1 <= month <= 12:
            return f"{text[:4]}-{month:02d}"
    raise ValueError("date cannot be normalized to YYYY-MM")


def _source(document: dict[str, Any], sheet: str, row: int, column: int | None = None) -> ProposalSource:
    cell_range = f"{get_column_letter(column)}{row}" if column else f"{row}:{row}"
    return ProposalSource(
        document_id=str(document.get("id", "")),
        filename=str(document.get("filename", "")),
        sheet=sheet,
        range=cell_range,
    )


def _row_value(row: tuple[Any, ...], one_based_index: int | None) -> Any:
    if one_based_index is None or one_based_index < 1 or one_based_index > len(row):
        return None
    return row[one_based_index - 1]


def _validate_request(request: IngestionToolRequest, worksheet: Any) -> None:
    contract = TOOL_CONTRACTS[request.tool]
    missing = [name for name in contract["required_columns"] if name not in request.columns]
    if missing:
        raise ValueError(f"{request.tool.value} missing column mappings: {', '.join(missing)}")
    if request.header_row > worksheet.max_row:
        raise ValueError("header_row is outside the worksheet")
    allowed = set(contract["required_columns"]) | set(contract["optional_columns"])
    unknown = sorted(set(request.columns) - allowed)
    if unknown:
        raise ValueError(f"unsupported canonical columns: {', '.join(unknown)}")
    header = next(
        worksheet.iter_rows(
            min_row=request.header_row,
            max_row=request.header_row,
            max_col=worksheet.max_column,
            values_only=True,
        ),
        (),
    )
    for canonical, index in request.columns.items():
        if index > worksheet.max_column:
            raise ValueError(f"column index for {canonical} is outside the worksheet")
        if _row_value(header, index) in (None, ""):
            raise ValueError(f"mapped header for {canonical} is empty")


def _cashbook(
    request: IngestionToolRequest,
    document: dict[str, Any],
    worksheet: Any,
) -> ToolExecutionResult:
    transactions: list[CashFlowTransaction] = []
    evidence: list[Evidence] = []
    warnings: list[str] = []
    rows = worksheet.iter_rows(
        min_row=request.header_row + 1,
        max_row=min(worksheet.max_row, MAX_DATA_ROWS),
        max_col=worksheet.max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, request.header_row + 1):
        raw_date = _row_value(row, request.columns["date"])
        raw_inflow = _row_value(row, request.columns["inflow"])
        raw_outflow = _row_value(row, request.columns["outflow"])
        if raw_date in (None, "") and raw_inflow in (None, "", 0) and raw_outflow in (None, "", 0):
            continue
        try:
            period = _period(raw_date)
            inflow = _decimal(raw_inflow, field="inflow")
            outflow = _decimal(raw_outflow, field="outflow")
        except ValueError as exc:
            warnings.append(f"Skipped {worksheet.title} row {row_number}: {exc}")
            continue
        optional = {
            key: _row_value(row, index)
            for key, index in request.columns.items()
            if key not in {"date", "inflow", "outflow"}
        }
        for direction, amount in ((CashDirection.INFLOW, inflow), (CashDirection.OUTFLOW, outflow)):
            if amount == 0:
                continue
            evidence_id = f"{document.get('id')}:{worksheet.title}:{row_number}:{direction.value}"
            transaction = CashFlowTransaction(
                period=period,
                date=raw_date if isinstance(raw_date, datetime | date) else None,
                direction=direction,
                activity=CashActivity.UNCLASSIFIED,
                category=_plain_text(optional.get("category") or optional.get("type") or "unclassified"),
                amount=amount,
                description=_plain_text(optional.get("description")) or None,
                source_ref=_plain_text(optional.get("source_ref") or optional.get("transaction_id")) or None,
                document_id=str(document.get("id", "")),
                filename=str(document.get("filename", "")),
                sheet=worksheet.title,
                row_number=row_number,
                evidence_id=evidence_id,
            )
            transactions.append(transaction)
            evidence.append(
                Evidence(
                    evidence_id=evidence_id,
                    source_type="uploaded_workbook",
                    title=f"{document.get('filename')} / {worksheet.title} row {row_number}",
                    document_id=str(document.get("id", "")),
                    reliability="high",
                )
            )
    return ToolExecutionResult(
        dataset=CashFlowDataset(transactions=transactions, source_type="ai_mapped_cashbook"),
        evidence=evidence,
        warnings=warnings,
    )


def _financial_facts(
    request: IngestionToolRequest,
    document: dict[str, Any],
    worksheet: Any,
) -> ToolExecutionResult:
    proposals: list[FieldProposal] = []
    dataset = CashFlowDataset(source_type="ai_mapped_facts")
    normalized_field_map = {_normalized_text(label): field for label, field in request.field_map.items()}
    header_values = worksheet.iter_rows(
        min_row=1,
        max_row=request.header_row,
        max_col=worksheet.max_column,
        values_only=True,
    )
    detected_currency = False
    detected_as_of = False
    for row_number, row in enumerate(header_values, 1):
        for column_number, raw_value in enumerate(row, 1):
            text = _plain_text(raw_value)
            if not text:
                continue
            currency_match = re.search(r"\b(VND|USD)\b", text, re.IGNORECASE)
            if currency_match and not detected_currency:
                proposals.append(
                    FieldProposal(
                        field="currency",
                        value=currency_match.group(1).upper(),
                        confidence="high",
                        sources=[_source(document, worksheet.title, row_number, column_number)],
                        generated_by_tool=request.tool.value,
                    )
                )
                detected_currency = True
            dates = re.findall(r"\b\d{4}-\d{2}-\d{2}\b", text)
            if len(dates) >= 2 and not detected_as_of:
                proposals.append(
                    FieldProposal(
                        field="cash_as_of",
                        value=dates[-1],
                        confidence="high",
                        sources=[_source(document, worksheet.title, row_number, column_number)],
                        generated_by_tool=request.tool.value,
                    )
                )
                detected_as_of = True
    rows = worksheet.iter_rows(
        min_row=request.header_row + 1,
        max_row=min(worksheet.max_row, MAX_DATA_ROWS),
        max_col=worksheet.max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, request.header_row + 1):
        label = _row_value(row, request.columns["label"])
        value = _row_value(row, request.columns["value"])
        if label in (None, "") or value in (None, ""):
            continue
        field = normalized_field_map.get(_normalized_text(label)) or FACT_ALIASES.get(_normalized_text(label))
        if field not in ALLOWED_FACT_FIELDS:
            continue
        parsed_value: Any = value
        if field not in {"currency", "cash_as_of"}:
            try:
                parsed_value = _decimal(value, field=field)
            except ValueError:
                continue
        if field == "opening_cash":
            dataset.opening_cash = parsed_value
        elif field in {"reported_ending_cash", "current_cash"}:
            dataset.reported_ending_cash = parsed_value
        proposals.append(
            FieldProposal(
                field=field,
                value=parsed_value,
                confidence="high",
                sources=[_source(document, worksheet.title, row_number, request.columns["value"])],
                generated_by_tool=request.tool.value,
            )
        )
        if field == "reported_ending_cash":
            proposals.append(
                FieldProposal(
                    field="current_cash",
                    value=parsed_value,
                    confidence="high",
                    sources=[_source(document, worksheet.title, row_number, request.columns["value"])],
                    generated_by_tool=request.tool.value,
                    warnings=["Mapped from a reported ending cash value; confirm the as-of date before applying."],
                )
            )
    return ToolExecutionResult(dataset=dataset, proposals=proposals)


def _sales(
    request: IngestionToolRequest,
    document: dict[str, Any],
    worksheet: Any,
) -> ToolExecutionResult:
    by_period: dict[str, Decimal] = defaultdict(Decimal)
    by_channel: dict[str, Decimal] = defaultdict(Decimal)
    by_payment: dict[str, Decimal] = defaultdict(Decimal)
    by_category: dict[str, Decimal] = defaultdict(Decimal)
    total = Decimal(0)
    quantity = Decimal(0)
    order_ids: set[str] = set()
    valid_rows = 0
    warnings: list[str] = []
    first_data_row: int | None = None
    last_data_row: int | None = None
    rows = worksheet.iter_rows(
        min_row=request.header_row + 1,
        max_row=min(worksheet.max_row, MAX_DATA_ROWS),
        max_col=worksheet.max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, request.header_row + 1):
        raw_date = _row_value(row, request.columns["date"])
        raw_net = _row_value(row, request.columns["net_amount"])
        if raw_date in (None, "") and raw_net in (None, ""):
            continue
        try:
            period = _period(raw_date)
            net = _decimal(raw_net, field="net_amount")
        except ValueError as exc:
            warnings.append(f"Skipped {worksheet.title} row {row_number}: {exc}")
            continue
        first_data_row = first_data_row or row_number
        last_data_row = row_number
        valid_rows += 1
        total += net
        by_period[period] += net
        if "quantity" in request.columns:
            try:
                quantity += _decimal(_row_value(row, request.columns["quantity"]), field="quantity")
            except ValueError:
                warnings.append(f"Invalid quantity ignored at {worksheet.title} row {row_number}.")
        for canonical, target in (
            ("channel", by_channel),
            ("payment_method", by_payment),
            ("category", by_category),
        ):
            if canonical in request.columns:
                label = _plain_text(_row_value(row, request.columns[canonical])) or "unknown"
                target[label] += net
        if "order_id" in request.columns:
            order_id = _plain_text(_row_value(row, request.columns["order_id"]))
            if order_id:
                order_ids.add(order_id)
    metrics: dict[str, Any] = {
        "net_sales": total,
        "quantity": quantity,
        "row_count": valid_rows,
        "by_period": dict(by_period),
        "by_channel": dict(by_channel),
        "by_payment_method": dict(by_payment),
        "by_category": dict(by_category),
        "order_count": len(order_ids) if order_ids else None,
        "average_order_value": total / len(order_ids) if order_ids else None,
    }
    if not order_ids:
        warnings.append("Average order value was not calculated because no order_id column was mapped.")
    source_range = None
    if first_data_row and last_data_row:
        source_range = f"{first_data_row}:{last_data_row}"
    proposal_source = ProposalSource(
        document_id=str(document.get("id", "")),
        filename=str(document.get("filename", "")),
        sheet=worksheet.title,
        range=source_range,
    )
    proposals = [
        FieldProposal(
            field="sales_support_metrics",
            value=metrics,
            confidence="high",
            sources=[proposal_source],
            generated_by_tool=request.tool.value,
            warnings=warnings,
        )
    ]
    return ToolExecutionResult(metrics={"sales": metrics}, proposals=proposals, warnings=warnings)


def _is_unpaid(value: Any) -> bool:
    text = _normalized_text(value)
    paid_markers = ("da thanh toan", "paid", "settled", "completed")
    unpaid_markers = ("chua thanh toan", "unpaid", "outstanding", "overdue", "partial")
    if any(marker in text for marker in unpaid_markers):
        return True
    return False if any(marker in text for marker in paid_markers) else False


def _purchases(
    request: IngestionToolRequest,
    document: dict[str, Any],
    worksheet: Any,
) -> ToolExecutionResult:
    by_period: dict[str, Decimal] = defaultdict(Decimal)
    by_category: dict[str, Decimal] = defaultdict(Decimal)
    by_supplier: dict[str, Decimal] = defaultdict(Decimal)
    total = Decimal(0)
    vat = Decimal(0)
    outstanding = Decimal(0)
    valid_rows = 0
    warnings: list[str] = []
    first_data_row: int | None = None
    last_data_row: int | None = None
    rows = worksheet.iter_rows(
        min_row=request.header_row + 1,
        max_row=min(worksheet.max_row, MAX_DATA_ROWS),
        max_col=worksheet.max_column,
        values_only=True,
    )
    for row_number, row in enumerate(rows, request.header_row + 1):
        raw_date = _row_value(row, request.columns["date"])
        raw_total = _row_value(row, request.columns["total_amount"])
        if raw_date in (None, "") and raw_total in (None, ""):
            continue
        try:
            period = _period(raw_date)
            amount = _decimal(raw_total, field="total_amount")
        except ValueError as exc:
            warnings.append(f"Skipped {worksheet.title} row {row_number}: {exc}")
            continue
        first_data_row = first_data_row or row_number
        last_data_row = row_number
        valid_rows += 1
        total += amount
        by_period[period] += amount
        if "category" in request.columns:
            category = _plain_text(_row_value(row, request.columns["category"])) or "unknown"
            by_category[category] += amount
        if "supplier" in request.columns:
            supplier = _plain_text(_row_value(row, request.columns["supplier"])) or "unknown"
            by_supplier[supplier] += amount
        if "vat_amount" in request.columns:
            try:
                vat += _decimal(_row_value(row, request.columns["vat_amount"]), field="vat_amount")
            except ValueError:
                warnings.append(f"Invalid VAT ignored at {worksheet.title} row {row_number}.")
        if "payment_status" in request.columns and _is_unpaid(_row_value(row, request.columns["payment_status"])):
            outstanding += amount
    metrics = {
        "total_purchases_and_expenses": total,
        "vat_amount": vat,
        "outstanding_payables": outstanding,
        "row_count": valid_rows,
        "by_period": dict(by_period),
        "by_category": dict(by_category),
        "by_supplier": dict(by_supplier),
    }
    source_range = None
    if first_data_row and last_data_row:
        source_range = f"{first_data_row}:{last_data_row}"
    source = ProposalSource(
        document_id=str(document.get("id", "")),
        filename=str(document.get("filename", "")),
        sheet=worksheet.title,
        range=source_range,
    )
    proposals = [
        FieldProposal(
            field="purchase_cost_metrics",
            value=metrics,
            confidence="high",
            sources=[source],
            generated_by_tool=request.tool.value,
        )
    ]
    if "payment_status" in request.columns:
        proposals.append(
            FieldProposal(
                field="accounts_payable",
                value=outstanding,
                confidence="medium",
                sources=[source],
                generated_by_tool=request.tool.value,
                warnings=["Derived only from mapped payment_status values in the imported purchase table."],
            )
        )
    return ToolExecutionResult(metrics={"purchases": metrics}, proposals=proposals, warnings=warnings)


def execute_ingestion_tool(
    request: IngestionToolRequest,
    documents_by_id: dict[str, dict[str, Any]],
) -> ToolExecutionResult:
    if request.document_id not in documents_by_id:
        raise ValueError("document_id is not available in this startup analysis")
    document = documents_by_id[request.document_id]
    path = Path(document.get("storage_path") or "")
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise ValueError("ingestion tools currently support stored .xlsx documents only")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if request.sheet not in workbook.sheetnames:
            raise ValueError("mapped worksheet does not exist")
        worksheet = workbook[request.sheet]
        if worksheet.max_row is None or worksheet.max_column is None:
            worksheet.calculate_dimension(force=True)
        _validate_request(request, worksheet)
        if request.tool == IngestionToolName.NORMALIZE_CASHBOOK:
            return _cashbook(request, document, worksheet)
        if request.tool == IngestionToolName.EXTRACT_FINANCIAL_FACTS:
            return _financial_facts(request, document, worksheet)
        if request.tool == IngestionToolName.SUMMARIZE_SALES:
            return _sales(request, document, worksheet)
        if request.tool == IngestionToolName.SUMMARIZE_PURCHASES:
            return _purchases(request, document, worksheet)
        raise ValueError(f"unsupported ingestion tool: {request.tool}")
    finally:
        workbook.close()
